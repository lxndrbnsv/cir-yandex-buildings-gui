import re
import os
from openpyxl import load_workbook, Workbook

RESULTS_HEADERS = [
    "Код", "Физлицо", "Водитель", "Группа", "Наименование",
    "Юридическое наименование", "Юридический адрес", "Фактический адрес",
    "Паспорт", "Должность", "Организация", "Область", "Район",
    "ИНН", "КПП", "ОКПО", "ОГРН", "ОКВЭД",
    "РабТел1", "РабТел2", "МобТел1", "МобТел2", "Факс", "Email",
    "Вебсайт", "БИК", "Банк", "К/с", "Р/с",
    "Скидка", "Заметки", "Тип организации",
]


def create_results_wb():
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws.append(RESULTS_HEADERS)
    wb.save("results.xlsx")


class ClearWB:
    def __init__(self):
        if os.path.exists("./results.xlsx"):
            os.remove("./results.xlsx")
        create_results_wb()


class CheckWB:
    def __init__(self):
        wb = load_workbook(filename="results.xlsx")
        ws = wb["Worksheet"]

        print(ws.max_row)


class NormalizePhoneNumber:
    def __init__(self, phone_string):
        stripped_symbols = (
            phone_string.replace("(", "")
            .replace(")", "")
            .replace(" ", "")
            .replace("-", "")
        )
        if stripped_symbols[0] == "8":
            stripped_list = list(stripped_symbols)
            stripped_list[0] = "+7"
            normalized = "".join(stripped_list)
        else:
            normalized = stripped_symbols

        self.normalize = normalized


def parse_query_string(query: str) -> str:
    address_items = query.split(",")

    address_data = [address_items[0]]

    for a in address_items[1:]:
        digits = re.findall(r"\d+", a)
        try:
            address_data.append(digits[0])
        except IndexError:
            pass

    return f"#{'-'.join(address_data)}"
