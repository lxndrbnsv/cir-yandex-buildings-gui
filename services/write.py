import os
from openpyxl import load_workbook
from services.util import create_results_wb


class WriteXLSX:
    def __init__(self, company_dict):
        if not os.path.exists("results.xlsx"):
            create_results_wb()
        wb = load_workbook(filename="results.xlsx")
        ws = wb["Worksheet"]

        line_to_write = ws.max_row + 1

        ws[f"D{line_to_write}"] = "Телемаркетинг"
        ws[f"E{line_to_write}"] = (
            f'{company_dict["name"]},' f' {company_dict["type"]}'
        )
        # ws[f"E{line_to_write}"] = f'{company_dict["address"]},'
        ws[f"H{line_to_write}"] = company_dict["address"]
        ws[f"S{line_to_write}"] = company_dict["phone"]
        ws[f"Y{line_to_write}"] = company_dict["website"]
        ws[f"AF{line_to_write}"] = company_dict["building_id"]

        wb.save(filename="results.xlsx")
